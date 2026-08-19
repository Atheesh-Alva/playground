import os
import csv
import datetime
import requests
import openpyxl
from zoneinfo import ZoneInfo

WEBHOOK_URL = os.environ["TEAMS_WEBHOOK_URL"]
EXCEL_PATH = "data/Standup_Updates.xlsx"
ARCHIVE_PATH = "data/updates_archive.csv"

def read_roster_and_updates():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # 1. Read Roster Sheet
    roster_sheet = wb["Roster"]
    roster = []
    for row in roster_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            roster.append({"order": int(row[0]), "name": row[1], "email": row[2]})
            
    # Sort roster by 'order' column just in case
    roster.sort(key=lambda x: x["order"])

    # 2. Read Today's Updates Sheet
    updates_sheet = wb["Today"]
    updates = {}
    for row in updates_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:  # row[1] is Email
            updates[row[1]] = row[2] if (len(row) > 2 and row[2]) else "_No update logged_"

    return roster, updates

def get_speaker_from_archive(roster):
    """Determines speaker by counting Mondays or looking at week numbers."""
    now_ist = datetime.datetime.now(ZoneInfo("Asia/Kolkata"))
    iso_week = now_ist.date().isocalendar()[1]
    
    # Continuous loop index based on ISO week number
    speaker_index = (iso_week - 1) % len(roster)
    return roster[speaker_index]

def archive_today_updates(today_str, roster, updates):
    """Appends today's updates to data/updates_archive.csv."""
    file_exists = os.path.exists(ARCHIVE_PATH)
    
    with open(ARCHIVE_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["date", "name", "update"])
            
        for member in roster:
            name = member["name"]
            email = member["email"]
            update_text = updates.get(email, "_No update logged_")
            writer.writerow([today_str, name, update_text])

def send_teams_card(speaker, roster, updates):
    facts = [
        {"title": member["name"], "value": updates.get(member["email"], "_No update logged_")}
        for member in roster
    ]

    card_payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "contentUrl": None,
                "content": {
                    "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                    "type": "AdaptiveCard",
                    "version": "1.4",
                    "body": [
                        {
                            "type": "TextBlock",
                            "text": "📢 Daily Standup Reminder (9:00 AM IST)",
                            "weight": "Bolder",
                            "size": "Large",
                            "color": "Accent"
                        },
                        {
                            "type": "Container",
                            "style": "warning",
                            "items": [
                                {
                                    "type": "TextBlock",
                                    "text": f"🎙️ **Weekly Speaker:** {speaker['name']} ({speaker['email']})",
                                    "weight": "Bolder",
                                    "wrap": True
                                }
                            ]
                        },
                        {
                            "type": "TextBlock",
                            "text": "📋 **Today's Team Updates**",
                            "weight": "Bolder",
                            "size": "Medium",
                            "spacing": "Medium"
                        },
                        {
                            "type": "FactSet",
                            "facts": facts
                        }
                    ]
                }
            }
        ]
    }

    resp = requests.post(WEBHOOK_URL, json=card_payload)
    resp.raise_for_status()

if __name__ == "__main__":
    now_ist = datetime.datetime.now(ZoneInfo("Asia/Kolkata"))
    today_str = now_ist.date().isoformat()
    
    roster, updates = read_roster_and_updates()
    speaker = get_speaker_from_archive(roster)
    
    send_teams_card(speaker, roster, updates)
    archive_today_updates(today_str, roster, updates)